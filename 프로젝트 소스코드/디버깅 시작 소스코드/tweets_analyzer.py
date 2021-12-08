#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Copyright (c) 2017 @x0rz
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, version 3 of the License.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# Usage:
# python tweets_analyzer.py -n screen_name
#
# Install:
# pip install tweepy ascii_graph tqdm numpy
"""from secrets import consumer_key, consumer_secret, access_token, access_token_secret"""
from __future__ import unicode_literals
from pandas import DataFrame
import tkinter.messagebox
from tkinter import scrolledtext
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter.font import Font
from typing import List

from ascii_graph import Pyasciigraph
from ascii_graph.colors import Gre, Yel, Red
from ascii_graph.colordata import hcolor
from numpy.core.fromnumeric import sort
from tqdm import tqdm_gui
from tkinter import *
from tkinter import simpledialog,messagebox
import tweepy
import numpy
import argparse
import collections
import datetime
import re
import json
import sys
import os
import copy
import openpyxl

from matplotlib.figure import Figure
import matplotlib.pyplot as plt
plt.rcParams["font.family"] = "Malgun Gothic"

__version__ = '0.2-dev'
from urllib.parse import urlparse
"""
try:
    from urllib.parse import urlparse
except ImportError:
    from urlparse import urlparse
"""

# Here are sglobals used to store data - I know it's dirty, whatever
start_date = 0
end_date = 0
export = ""
jsono = {}
tweet_limit = 1000  # default 트윗 수
#ansi_escape = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')

activity_hourly = {
    ("%2i:00" % i).replace(" ", "0"): 0 for i in range(24)
}

activity_weekly = {
    "%i" % i: 0 for i in range(7)
}

detected_langs = collections.Counter()
detected_sources = collections.Counter()
detected_places = collections.Counter()
geo_enabled_tweets = 0
detected_hashtags = collections.Counter()
detected_domains = collections.Counter()
detected_timezones = collections.Counter()
retweets = 0
retweeted_users = collections.Counter()
mentioned_users = collections.Counter()
id_screen_names = {}
frienmns_timezone = collections.Counter()
friends_lang = collections.Counter()


def process_tweet(tweet):
    """ Processing a single Tweet and updating our datasets """
    global start_date
    global end_date
    global geo_enabled_tweets
    global retweets

    tw_date = tweet.created_at

    # Updating most recent tweet
    end_date = end_date or tw_date
    start_date = tw_date

    # Handling retweets
    try:
        # We use id to get unique accounts (screen_name can be changed)
        rt_id_user = tweet.retweeted_status.user.id_str
        retweeted_users[rt_id_user] += 1

        if tweet.retweeted_status.user.screen_name not in id_screen_names:
            id_screen_names[rt_id_user] = "@%s" % tweet.retweeted_status.user.screen_name

        retweets += 1
    except:
        pass

    # Adding timezone from profile offset to set to local hours

    # Updating our activity datasets (distribution maps)
    activity_hourly["%s:00" % str(tw_date.hour).zfill(2)] += 1
    activity_weekly[str(tw_date.weekday())] += 1

    # Updating langs
    detected_langs[tweet.lang] += 1

    # Updating sources
    detected_sources[tweet.source] += 1

    # Detecting geolocation
    if tweet.place:
        geo_enabled_tweets += 1
        tweet.place.name = tweet.place.name
        detected_places[tweet.place.name] += 1

    # Updating hashtags list
    if tweet.entities['hashtags']:
        for ht in tweet.entities['hashtags']:
            ht['text'] = "#%s" % ht['text']
            detected_hashtags[ht['text']] += 1

    # Updating domains list
    if tweet.entities['urls']:
        for url in tweet.entities['urls']:
            domain = urlparse(url['expanded_url']).netloc
            # removing twitter.com from domains (not very relevant)
            if domain != "twitter.com":
                detected_domains[domain] += 1

    # Updating mentioned users list
    if tweet.entities['user_mentions']:
        for ht in tweet.entities['user_mentions']:
            mentioned_users[ht['id_str']] += 1
            if not ht['screen_name'] in id_screen_names:
                id_screen_names[ht['id_str']] = "@%s" % ht['screen_name']


def get_tweets(api, username, fh, limit):
    """ Download Tweets from username account """
    global geo_enabled_tweets
    global retweets
    retweets = 0
    geo_enabled_tweets = 0
    detected_langs.clear()
    detected_sources.clear()
    detected_places.clear()
    detected_hashtags.clear()
    detected_domains.clear()
    retweeted_users.clear()
    mentioned_users.clear()
    #friends_timezone.clear()
    detected_timezones.clear()
    friends_lang.clear()
    id_screen_names = {}
    for status in tqdm_gui(tweepy.Cursor(api.user_timeline, screen_name=username).items(limit), unit="tw", total=limit):
            process_tweet(status)
    

def int_to_weekday(day):
    weekdays = "Monday Tuesday Wednesday Thursday Friday Saturday Sunday".split()
    return weekdays[int(day) % len(weekdays)]



def cprint(strng):
    print(strng)
    #export_string(strng)
    

def main():
    #global color_supported
    #color_supported = supports_color()
   
    consumer_key = "4jFe7U2QXf8qeaybYzffuLr5h"
    consumer_secret = "KxHepwVJUSxDSM4V6yueTWM9WAK5IVAm0Fbn6OfEieCGuEVr1J"

    access_token = "1456829819967406082-ZHXxp0qD3K7NU2qyUI4IbyLczj75rC"
    access_token_secret = "ojE68aAwLjb4gOOAkMjjJMWTjMaGo8ar0ykffZ9QIczm6"

    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_token, access_token_secret)
    twitter_api = tweepy.API(auth)

    now = datetime.datetime.now()
    #save_path = save_folder+"/"+args.name
    save_file = False
    
    # Getting general account's metadata
    #cprint("[+] Getting @%s account data..." % args.name)

    #lab1 = Label(win)

    
    retweeted_users_names = {}
    mentioned_users_names = {}
    
    win = Tk()
    win.geometry("1040x765")
    win.title("트위터 분석기")
    win.resizable(False, False)

    font = Font(win, weight="bold")

    def get_act():
        reset_canvas()
        change_button_color()
        btn_get_activity.config(bg="lightgrey")
    
        sum = numpy.sum(list(detected_hashtags.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="해시태그 TOP 10", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            sorted_keys = sorted(
                detected_hashtags, key=detected_hashtags.get, reverse=True)
            max_len_key = max([len(x) for x in sorted_keys][:10])
            sorted_values = []
            a = 1
            for k in sorted_keys:
                txt.insert(END, ("{0}. {1}\t\t\t\t{2} ({3}%)").format(
                    a, k, detected_hashtags[k], round(((float(detected_hashtags[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                sorted_values.append(detected_hashtags[k])
                a += 1
                i += 1
                if(i >= 10):
                    break

            x = numpy.array(sorted_keys[:10])
            y = numpy.array(sorted_values)
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_daily_timeline():
        reset_canvas()
        change_button_color()
        btn_get_daily_timeline.config(bg="lightgrey")
        sum = numpy.sum(list(activity_hourly.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="시간대별 타임라인", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            max_value = max(activity_hourly.values())
            for k in activity_hourly:
                if(activity_hourly[k] != 0):
                    block_count = max_value/activity_hourly[k]
                    block = ""
                    num = int(10*(activity_hourly[k]/max_value)+1)
                    for count in range(1, num):
                        block += "■"
                else:
                    block = ""
                txt.insert(END, (" {0}\t\t\t{1}\t{2}\t({3}%)").format(
                    block, activity_hourly[k], k, round(((float(activity_hourly[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                i += 1
                if(i >= 24):
                    break

            x = numpy.array(list(activity_hourly.keys()))
            y = numpy.array(list(activity_hourly.values()))
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_weekly_timeline():
        reset_canvas()
        change_button_color()
        btn_get_weekly_timeline.config(bg="lightgrey")

        sum = numpy.sum(list(activity_weekly.values()))
        i = 0
        date = ["Monday", "Tuesday", "Wednesday",
                "Thursday", "Friday", "Saturday", "Sunday"]

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="요일별 타임라인", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            max_value = max(activity_weekly.values())

            date_list = []
            for k in activity_weekly:
                date_list.append(date[int(k)])

            for k in activity_weekly:
                if(activity_weekly[k] != 0):
                    block_count = max_value/activity_weekly[k]
                    block = ""
                    num = int(10*(activity_weekly[k]/max_value)+1)
                    for count in range(1, num):
                        block += "■"
                else:
                    block = ""
                txt.insert(END, (" {0}\t\t\t{1}\t{2}\t\t({3}%)").format(
                    block, activity_weekly[k], date[int(k)], round(((float(activity_weekly[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                i += 1
                if(i >= 7):
                    break

            x = numpy.array(date_list)
            y = numpy.array(list(activity_weekly.values()))
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_lang():
        reset_canvas()
        change_button_color()
        btn_get_language.config(bg="lightgrey")

        sum = numpy.sum(list(detected_langs.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="사용 언어 TOP 10", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            sorted_keys = sorted(
                detected_langs, key=detected_langs.get, reverse=True)
            max_len_key = max([len(x) for x in sorted_keys][:10])
            sorted_values = []
            a = 1
            for k in sorted_keys:
                txt.insert(END, ("{0}. {1}\t\t\t\t{2} ({3}%)").format(
                    a, k, detected_langs[k], round(((float(detected_langs[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                sorted_values.append(detected_langs[k])
                a += 1
                i += 1
                if(i >= 10):
                    break

            x = numpy.array(sorted_keys[:10])
            y = numpy.array(sorted_values)
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0,columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_source():
        reset_canvas()
        change_button_color()
        btn_get_source.config(bg="lightgrey")

        sum = numpy.sum(list(detected_sources.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="사용 소스 TOP 10", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            sorted_keys = sorted(
                detected_sources, key=detected_sources.get, reverse=True)
            max_len_key = max([len(x) for x in sorted_keys][:10])
            sorted_values = []
            a = 1
            for k in sorted_keys:
                txt.insert(END, ("{0}. {1}\t\t\t\t{2} ({3}%)").format(
                    a, k, detected_sources[k], round(((float(detected_sources[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                sorted_values.append(detected_sources[k])
                a += 1
                i += 1
                if(i >= 10):
                    break

            x = numpy.array(sorted_keys[:10])
            y = numpy.array(sorted_values)
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_geo():
        txt.delete("1.0", END)
        txt.insert(END, "asd")
        
        
    def get_retw():
        reset_canvas()
        change_button_color()
        btn_get_retweets.config(bg="lightgrey")
        
        sum = numpy.sum(list(retweeted_users_names.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="리트윗 사용자 TOP 10", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            sorted_keys = sorted(retweeted_users_names,
                                 key=retweeted_users_names.get, reverse=True)
            max_len_key = max([len(x) for x in sorted_keys][:10])
            sorted_values = []
            a = 1
            for k in sorted_keys:
                txt.insert(END, ("{0}. {1}\t\t\t\t{2} ({3}%)").format(
                    a, k, retweeted_users_names[k], round(((float(retweeted_users_names[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                sorted_values.append(retweeted_users_names[k])
                a += 1
                i += 1
                if(i >= 10):
                    break

            x = numpy.array(sorted_keys[:10])
            y = numpy.array(sorted_values)
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def get_domains():
        reset_canvas()
        change_button_color()
        btn_get_domains.config(bg="lightgrey")
       
        sum = numpy.sum(list(detected_domains.values()))
        i = 0

        txt.delete("1.0", END)

        if(sum):
            label = Label(win, text="참조한 사이트 TOP 10", height=2, font=font)
            label.grid(row=1, column=0, columnspan=10, sticky=N+E+W+S)

            sorted_keys = sorted(
                detected_domains, key=detected_domains.get, reverse=True)
            max_len_key = max([len(x) for x in sorted_keys][:10])
            sorted_values = []
            a = 1
            for k in sorted_keys:
                txt.insert(END, ("{0}. {1}\t\t\t\t{2} ({3}%)").format(
                    a, k, detected_domains[k], round(((float(detected_domains[k] / sum) * 100)), 2)))
                txt.insert(END, "\n")
                sorted_values.append(detected_domains[k])
                a += 1
                i += 1
                if(i >= 10):
                    break

            x = numpy.array(sorted_keys[:10])
            y = numpy.array(sorted_values)
            fig.add_subplot(1, 1, 1).bar(x, y)

            canvas.draw()
            canvas.get_tk_widget().grid(row=2, column=0, columnspan=10, sticky=N+E+W+S)

            txt.grid(row=3, column=0, columnspan=10, sticky=N+E+W+S)
        else:
            tkinter.messagebox.showwarning("No Data", "데이터가 없습니다.")

    def out_excelFile():  # 엑셀파일로 데이터 출력
        dir = jsono['user_name'] + '_Data.xlsx'
        excel = openpyxl.Workbook()

        excel_ws = excel.active  # 현재 활성화된 시트 선택
        # (방금 만들었으니 시트가 1개밖에 없음 -> 자동으로 1개 활성화)
        excel_ws.cell(row=2, column=1).value = 'Cheese'

        sum = numpy.sum(list(detected_hashtags.values()))
        i = 0

        excel_ws.cell(row=1, column=1).value = "detected_hashtags"
        if(sum):
            sorted_keys = sorted(
                detected_hashtags, key=detected_hashtags.get, reverse=True)
            for k in sorted_keys:
                excel_ws.cell(row=2, column=i+1).value = k
                excel_ws.cell(row=3, column=i+1).value = detected_hashtags[k]
                excel_ws.cell(row=4, column=i+1).value = round(
                    ((float(detected_hashtags[k] / sum) * 100)), 2).__str__() + '%'
                i += 1

        sum = numpy.sum(list(activity_hourly.values()))
        i = 0

        excel_ws.cell(row=5, column=1).value = "activity_hourly"
        if(sum):
            max_value = max(activity_hourly.values())
            for k in activity_hourly:
                excel_ws.cell(row=6, column=i+1).value = k
                excel_ws.cell(row=7, column=i+1).value = activity_hourly[k]
                excel_ws.cell(row=8, column=i+1).value = round(
                    ((float(activity_hourly[k] / sum) * 100)), 2).__str__() + '%'
                i += 1
                if(i >= 24):
                    break

        sum = numpy.sum(list(activity_weekly.values()))
        i = 0
        date = ["Monday", "Tuesday", "Wednesday",
                "Thursday", "Friday", "Saturday", "Sunday"]

        excel_ws.cell(row=9, column=1).value = "activity_weekly"
        if(sum):
            for k in activity_weekly:
                excel_ws.cell(row=10, column=i+1).value = date[int(k)]
                excel_ws.cell(row=11, column=i+1).value = activity_weekly[k]
                excel_ws.cell(row=12, column=i+1).value = round(
                    ((float(activity_weekly[k] / sum) * 100)), 2).__str__() + '%'
                i += 1
                if(i >= 7):
                    break

        sum = numpy.sum(list(detected_langs.values()))
        i = 0

        excel_ws.cell(row=13, column=1).value = "detected_langs"
        if(sum):
            sorted_keys = sorted(
                detected_langs, key=detected_langs.get, reverse=True)
            for k in sorted_keys:
                excel_ws.cell(row=14, column=i+1).value = k
                excel_ws.cell(row=15, column=i+1).value = detected_langs[k]
                excel_ws.cell(row=16, column=i+1).value = round(
                    ((float(detected_langs[k] / sum) * 100)), 2).__str__() + '%'
                i += 1

        sum = numpy.sum(list(detected_sources.values()))
        i = 0

        excel_ws.cell(row=17, column=1).value = "detected_sources"
        if(sum):
            sorted_keys = sorted(
                detected_sources, key=detected_sources.get, reverse=True)
            for k in sorted_keys:
                excel_ws.cell(row=18, column=i+1).value = k
                excel_ws.cell(row=19, column=i+1).value = detected_sources[k]
                excel_ws.cell(row=20, column=i+1).value = round(
                    ((float(detected_sources[k] / sum) * 100)), 2).__str__() + '%'
                i += 1

        sum = numpy.sum(list(retweeted_users_names.values()))
        i = 0

        excel_ws.cell(row=21, column=1).value = "retweeted_users_names"

        if(sum):
            sorted_keys = sorted(retweeted_users_names,
                                 key=retweeted_users_names.get, reverse=True)
            for k in sorted_keys:
                excel_ws.cell(row=22, column=i+1).value = k
                excel_ws.cell(row=23, column=i +
                              1).value = retweeted_users_names[k]
                excel_ws.cell(row=24, column=i+1).value = round(
                    ((float(retweeted_users_names[k] / sum) * 100)), 2).__str__() + '%'
                i += 1

        sum = numpy.sum(list(detected_domains.values()))
        i = 0

        excel_ws.cell(row=21, column=1).value = "detected_domains"
        if(sum):
            sorted_keys = sorted(
                detected_domains, key=detected_domains.get, reverse=True)
            
            for k in sorted_keys:
                excel_ws.cell(row=22, column=i+1).value = k
                excel_ws.cell(row=23, column=i+1).value = detected_domains[k]
                excel_ws.cell(row=24, column=i+1).value = round(
                    ((float(detected_domains[k] / sum) * 100)), 2).__str__() + '%'
                i += 1

        
        excel.save(dir)  # 절대 경로로 하지 않으면 현재 디렉토리에 그냥 파일이 생긴다.
        messagebox.showinfo("성공","엑셀 데이터 출력 성공")

    def change_button_color():
        btn_get_activity.config(bg="SystemButtonFace")
        btn_get_daily_timeline.config(bg="SystemButtonFace")
        btn_get_weekly_timeline.config(bg="SystemButtonFace")
        btn_get_language.config(bg="SystemButtonFace")
        btn_get_source.config(bg="SystemButtonFace")
        btn_get_retweets.config(bg="SystemButtonFace")
        btn_get_domains.config(bg="SystemButtonFace")

    def reset_canvas():
        txt.grid_remove()
        fig.clear()
        canvas.draw()

    def search_tweet(self):
        try: 
            get_tweets(twitter_api, search.get(), save_file, tweet_limit)
            #mentioned_users_names = {}
            jsono['user_name'] = search.get()
            for k in mentioned_users.keys():
                mentioned_users_names[id_screen_names[k]] = mentioned_users[k]
            #retweeted_users_names = {}
            for k in retweeted_users.keys():
                retweeted_users_names[id_screen_names[k]] = retweeted_users[k]
        except tweepy.error.TweepError as e:
            messagebox.showerror("오류", "접근할 수 없습니다.")
        
    
    def set_limit():
        global tweet_limit
        while(1):
            limit = simpledialog.askinteger(title="트윗", prompt="트윗 수를 입력하세요.")
            if(limit > 0):
                tweet_limit = limit
                break;
            else:
                messagebox.showwarning("잘못된 값", "음수입니다")
    
    fig = Figure(figsize=(13, 7), dpi=80)  # 13 11로 바꿈
    canvas = FigureCanvasTkAgg(fig, master=win)

    btn_get_activity = Button(win, text="활동내역", command=get_act)
   
    
    btn_get_daily_timeline = Button(
        win, text="타임라인(시간대별)", command=get_daily_timeline)
   
    
    btn_get_weekly_timeline = Button(
        win, text="타임라인(요일별)", command=get_weekly_timeline)
   
    
    btn_get_language = Button(win, text="사용언어", command=get_lang)
    
    
    btn_get_source = Button(win, text="사용소스", command=get_source)
    
    
    btn_get_retweets = Button(win, text="리트윗내역", command=get_retw)
   
    
    btn_get_domains = Button(win, text="도메인참조", command=get_domains)
   

    btn_make_excelFile = Button(win, text="엑셀출력", command=out_excelFile)

   #검색창
    search_label = Label(win, text="스크린")
    search = Entry(win)
    search.bind("<Return>", search_tweet)

    btn_limit = Button(win, text="트윗 수", command=set_limit)
    
    btn_get_activity.grid(row=0, column=0, sticky=N+E+W+S)
    btn_get_daily_timeline.grid(row=0, column=1, sticky=N+E+W+S)
    btn_get_weekly_timeline.grid(row=0, column=2, sticky=N+E+W+S)
    btn_get_language.grid(row=0, column=3, sticky=N+E+W+S)
    btn_get_source.grid(row=0, column=4, sticky=N+E+W+S)
    btn_get_retweets.grid(row=0, column=5, sticky=N+E+W+S)
    btn_get_domains.grid(row=0, column=6, sticky=N+E+W+S)

    btn_make_excelFile.grid(row=0, column=7, sticky=N+E+W+S)
    
    btn_limit.grid(row=0, column=8, sticky=N+E+W+S)
    search.grid(row=0, column=9, sticky=N+E+W+S)
    txt = scrolledtext.ScrolledText(win, width=30, height=10)

    win.mainloop()


if __name__ == '__main__':
    try:
        main()
    except tweepy.error.TweepError as e:
        cprint("[\033[91m!\033[0m] Twitter error: %s" % e)
    except Exception as e:
        cprint("[\033[91m!\033[0m] Error: %s" % e)
